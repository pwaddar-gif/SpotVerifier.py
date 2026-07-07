import sys
import os
import pandas as pd
from datetime import datetime
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QLabel, QPushButton, QFileDialog, 
                             QTableWidget, QTableWidgetItem, QTabWidget, 
                             QLineEdit, QHeaderView, QMessageBox, QDropEvent,
                             QProgressBar)
from PyQt6.QtCore import Qt, QMimeData
from PyQt6.QtGui import QColor, QFont, QDragEnterEvent
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

class BroadcastSpotVerifier(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("🎬 Broadcast Spot Verification Tool - One Click")
        self.setGeometry(100, 100, 1300, 900)
        self.setAcceptDrops(True)
        
        self.spot_sheet_path = ""
        self.as_run_path = ""
        self.df_spot = None
        self.df_asrun = None
        self.results = {
            "played": [],
            "missed": [],
            "extra": [],
            "duration_mismatch": [],
            "wrong_order": [],
            "duplicate": []
        }
        
        self.init_ui()

    def init_ui(self):
        # Dark Professional Theme
        self.setStyleSheet("""
            QMainWindow { background-color: #0D1B2A; }
            QLabel { color: #E0E0E0; font-size: 12px; }
            QPushButton { background-color: #1F3A52; color: #FFFFFF; border: 2px solid #00A8E8; 
                          padding: 10px 20px; border-radius: 5px; font-weight: bold; font-size: 12px; }
            QPushButton:hover { background-color: #00A8E8; color: #0D1B2A; }
            QPushButton:pressed { background-color: #0077B6; }
            QLineEdit { background-color: #1E3A52; color: #FFFFFF; border: 2px solid #444; 
                        padding: 8px; border-radius: 4px; }
            QTableWidget { background-color: #1A2F42; color: #E0E0E0; gridline-color: #333; 
                          border: 1px solid #333; }
            QHeaderView::section { background-color: #0F2640; color: #00D4FF; padding: 8px; 
                                  border: 1px solid #333; font-weight: bold; }
            QTabWidget::pane { border: 1px solid #333; background-color: #1A2F42; }
            QTabBar::tab { background-color: #1F3A52; color: #A0A0A0; padding: 10px 20px; 
                          border: 1px solid #333; }
            QTabBar::tab:selected { background-color: #0F2640; color: #00D4FF; 
                                   border-bottom: 3px solid #00D4FF; }
            QProgressBar { background-color: #1E3A52; border: 1px solid #333; border-radius: 4px; }
            QProgressBar::chunk { background-color: #00A8E8; }
        """)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(15)
        main_layout.setContentsMargins(20, 20, 20, 20)

        # ---- TITLE ----
        title = QLabel("🎬 Broadcast Spot Verification - One Click")
        title.setFont(QFont("Arial", 16, QFont.Weight.Bold))
        title.setStyleSheet("color: #00D4FF;")
        main_layout.addWidget(title)

        # ---- FILE SELECTION SECTION ----
        file_frame = QWidget()
        file_frame.setStyleSheet("background-color: #1A2F42; border: 2px dashed #00A8E8; border-radius: 8px; padding: 15px;")
        file_layout = QVBoxLayout(file_frame)

        # Spot Sheet
        spot_row = QHBoxLayout()
        self.lbl_spot = QLabel("📊 Spot Sheet: Not Selected (Drag & Drop or Browse)")
        self.lbl_spot.setStyleSheet("color: #FFB74D; font-weight: bold;")
        btn_spot = QPushButton("📂 Browse Spot Sheet")
        btn_spot.clicked.connect(self.browse_spot_sheet)
        btn_spot.setMaximumWidth(200)
        spot_row.addWidget(btn_spot)
        spot_row.addWidget(self.lbl_spot, 1)
        file_layout.addLayout(spot_row)

        # AsRun Log
        asrun_row = QHBoxLayout()
        self.lbl_asrun = QLabel("📋 AsRun Log: Not Selected (Drag & Drop or Browse)")
        self.lbl_asrun.setStyleSheet("color: #FFB74D; font-weight: bold;")
        btn_asrun = QPushButton("📂 Browse AsRun Log")
        btn_asrun.clicked.connect(self.browse_asrun_log)
        btn_asrun.setMaximumWidth(200)
        asrun_row.addWidget(btn_asrun)
        asrun_row.addWidget(self.lbl_asrun, 1)
        file_layout.addLayout(asrun_row)

        main_layout.addWidget(file_frame)

        # ---- VERIFY BUTTON (ONE-CLICK) ----
        self.btn_verify = QPushButton("✅ VERIFY SPOTS (ONE-CLICK)")
        self.btn_verify.setStyleSheet("""
            QPushButton {
                background-color: #00A8E8; 
                color: #0D1B2A; 
                font-size: 16px; 
                padding: 15px; 
                margin: 10px 0px; 
                font-weight: bold;
                border: none;
            }
            QPushButton:hover { background-color: #00D4FF; }
            QPushButton:pressed { background-color: #0077B6; }
        """)
        self.btn_verify.clicked.connect(self.verify_logs)
        main_layout.addWidget(self.btn_verify)

        # ---- PROGRESS BAR ----
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        self.progress.setMaximum(100)
        self.progress.setStyleSheet("QProgressBar { height: 10px; }")
        main_layout.addWidget(self.progress)

        # ---- SUMMARY DASHBOARD ----
        self.summary_widget = QWidget()
        self.summary_widget.setStyleSheet("""
            background-color: #0F2640; 
            border: 2px solid #00A8E8; 
            border-radius: 8px; 
            padding: 15px;
        """)
        sum_layout = QHBoxLayout(self.summary_widget)
        sum_layout.setSpacing(20)

        self.lbl_scheduled = QLabel("📅 Scheduled: --")
        self.lbl_played = QLabel("🟢 Played: --")
        self.lbl_missed = QLabel("🔴 Missed: --")
        self.lbl_extra = QLabel("🔵 Extra: --")
        self.lbl_mismatch = QLabel("🟡 Mismatch: --")
        self.lbl_status = QLabel("⏳ Status: Ready")

        for lbl in [self.lbl_scheduled, self.lbl_played, self.lbl_missed, self.lbl_extra, self.lbl_mismatch, self.lbl_status]:
            lbl.setFont(QFont("Arial", 11, QFont.Weight.Bold))
            lbl.setStyleSheet("color: #00D4FF;")
            sum_layout.addWidget(lbl)

        main_layout.addWidget(self.summary_widget)

        # ---- SEARCH BAR ----
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("🔍 Quick Filter:"))
        self.txt_search = QLineEdit()
        self.txt_search.setPlaceholderText("Search by Creative ID, Break, Brand...")
        self.txt_search.textChanged.connect(self.filter_tables)
        self.txt_search.setMaximumWidth(400)
        search_layout.addWidget(self.txt_search)
        search_layout.addStretch()
        main_layout.addLayout(search_layout)

        # ---- TABS FOR RESULTS ----
        self.tabs = QTabWidget()
        self.tbl_played = QTableWidget()
        self.tbl_missed = QTableWidget()
        self.tbl_extra = QTableWidget()
        self.tbl_mismatch = QTableWidget()

        self.tabs.addTab(self.tbl_played, "🟢 Played Correctly")
        self.tabs.addTab(self.tbl_missed, "🔴 Missed Spots")
        self.tabs.addTab(self.tbl_extra, "🔵 Extra/Unscheduled")
        self.tabs.addTab(self.tbl_mismatch, "🟡 Duration Mismatch")

        main_layout.addWidget(self.tabs)

        # ---- EXPORT BUTTONS ----
        export_layout = QHBoxLayout()
        export_layout.addStretch()
        btn_excel = QPushButton("📄 Export to Excel (Color Coded)")
        btn_excel.clicked.connect(lambda: self.export_data("xlsx"))
        btn_csv = QPushButton("📊 Export to CSV")
        btn_csv.clicked.connect(lambda: self.export_data("csv"))
        export_layout.addWidget(btn_excel)
        export_layout.addWidget(btn_csv)
        main_layout.addLayout(export_layout)

    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event: QDropEvent):
        files = [u.toLocalFile() for u in event.mimeData().urls()]
        for file in files:
            if file.endswith('.csv'):
                if not self.spot_sheet_path:
                    self.spot_sheet_path = file
                    self.lbl_spot.setText(f"✓ {os.path.basename(file)}")
                elif not self.as_run_path:
                    self.as_run_path = file
                    self.lbl_asrun.setText(f"✓ {os.path.basename(file)}")

    def browse_spot_sheet(self):
        file, _ = QFileDialog.getOpenFileName(self, "Select Spot Sheet CSV", "", "CSV Files (*.csv)")
        if file:
            self.spot_sheet_path = file
            self.lbl_spot.setText(f"✓ {os.path.basename(file)}")

    def browse_asrun_log(self):
        file, _ = QFileDialog.getOpenFileName(self, "Select AsRun Log CSV", "", "CSV Files (*.csv)")
        if file:
            self.as_run_path = file
            self.lbl_asrun.setText(f"✓ {os.path.basename(file)}")

    def verify_logs(self):
        if not self.spot_sheet_path or not self.as_run_path:
            QMessageBox.warning(self, "❌ Error", "Please select both Spot Sheet and AsRun Log files.")
            return

        self.progress.setVisible(True)
        self.progress.setValue(0)

        try:
            self.lbl_status.setText("⏳ Status: Loading files...")
            
            # Load files
            self.df_spot = pd.read_csv(self.spot_sheet_path).fillna("Unknown")
            self.df_asrun = pd.read_csv(self.as_run_path).fillna("Unknown")
            self.progress.setValue(20)

            self.lbl_status.setText("⏳ Status: Mapping columns...")
            spot_map = self.detect_columns(self.df_spot, "spot")
            asrun_map = self.detect_columns(self.df_asrun, "asrun")
            self.progress.setValue(40)

            # Clear previous results
            self.results = {
                "played": [],
                "missed": [],
                "extra": [],
                "duration_mismatch": [],
                "wrong_order": [],
                "duplicate": []
            }

            self.lbl_status.setText("⏳ Status: Verifying spots...")
            asrun_played_indices = set()

            # Process scheduled spots
            for idx, s_row in self.df_spot.iterrows():
                c_id = str(s_row[spot_map["id"]]).strip()
                brk = str(s_row[spot_map["break"]]).strip() if "break" in spot_map else "--"
                sched_dur = str(s_row[spot_map["duration"]]).strip() if "duration" in spot_map else "--"
                brand = str(s_row[spot_map["brand"]]).strip() if "brand" in spot_map else "--"
                slot = str(s_row[spot_map["slot"]]).strip() if "slot" in spot_map else "--"

                # Match in AsRun Log
                matched = self.df_asrun[
                    (self.df_asrun[asrun_map["id"]].astype(str).str.strip() == c_id) &
                    (~self.df_asrun.index.isin(asrun_played_indices))
                ]

                if not matched.empty:
                    match_idx = matched.index[0]
                    asrun_played_indices.add(match_idx)
                    a_row = matched.iloc[0]
                    actual_dur = str(a_row[asrun_map["duration"]]).strip() if "duration" in asrun_map else "--"
                    actual_time = str(a_row[asrun_map["time"]]).strip() if "time" in asrun_map else "--"

                    record = [c_id, brk, slot, sched_dur, actual_dur, brand, actual_time]
                    if sched_dur == actual_dur or sched_dur == "--" or actual_dur == "--":
                        self.results["played"].append(record)
                    else:
                        self.results["duration_mismatch"].append(record)
                else:
                    record = [c_id, brk, slot, sched_dur, "--", brand, "--"]
                    self.results["missed"].append(record)

            self.progress.setValue(70)

            # Unmatched entries in AsRun are extra spots
            for idx, a_row in self.df_asrun.iterrows():
                if idx not in asrun_played_indices:
                    c_id = str(a_row[asrun_map["id"]]).strip()
                    brk = str(a_row[asrun_map["break"]]).strip() if "break" in asrun_map else "--"
                    actual_dur = str(a_row[asrun_map["duration"]]).strip() if "duration" in asrun_map else "--"
                    actual_time = str(a_row[asrun_map["time"]]).strip() if "time" in asrun_map else "--"
                    record = [c_id, brk, "--", "--", actual_dur, "--", actual_time]
                    self.results["extra"].append(record)

            self.progress.setValue(85)

            # Update Dashboard
            self.lbl_scheduled.setText(f"📅 Scheduled: {len(self.df_spot)}")
            self.lbl_played.setText(f"🟢 Played: {len(self.results['played'])}")
            self.lbl_missed.setText(f"🔴 Missed: {len(self.results['missed'])}")
            self.lbl_extra.setText(f"🔵 Extra: {len(self.results['extra'])}")
            self.lbl_mismatch.setText(f"🟡 Mismatch: {len(self.results['duration_mismatch'])}")

            # Update Tables
            headers = ["Creative ID", "Break", "Slot", "Sched Dur", "Actual Dur", "Brand", "Air Time"]
            self.populate_table(self.tbl_played, self.results["played"], headers, QColor("#1B5E20"))
            self.populate_table(self.tbl_missed, self.results["missed"], headers, QColor("#B71C1C"))
            self.populate_table(self.tbl_extra, self.results["extra"], headers, QColor("#0D47A1"))
            self.populate_table(self.tbl_mismatch, self.results["duration_mismatch"], headers, QColor("#F57F17"))

            self.progress.setValue(100)
            self.lbl_status.setText(f"✅ Status: Verification Complete! ({datetime.now().strftime('%H:%M:%S')})")
            QMessageBox.information(self, "✅ Success", "Verification Complete!\n\nReview results in tabs above.")

        except Exception as e:
            self.lbl_status.setText("❌ Status: Error")
            QMessageBox.critical(self, "❌ Error", f"Error during verification:\n{str(e)}")
        finally:
            self.progress.setVisible(False)

    def detect_columns(self, df, file_type="spot"):
        """Intelligent mapping for broadcast formats"""
        mapping = {}
        cols_lower = [str(c).strip().lower() for c in df.columns]

        for i, col in enumerate(df.columns):
            c_low = cols_lower[i]
            
            if file_type == "spot":
                if "creative" in c_low or "blaze" in c_low or "ad id" in c_low:
                    mapping["id"] = col
                elif "break" in c_low:
                    mapping["break"] = col
                elif "slot" in c_low or "position" in c_low:
                    mapping["slot"] = col
                elif "duration" in c_low or "dur" in c_low:
                    mapping["duration"] = col
                elif "brand" in c_low:
                    mapping["brand"] = col
            else:  # asrun
                if "creative" in c_low or "ad id" in c_low:
                    mapping["id"] = col
                elif "break" in c_low:
                    mapping["break"] = col
                elif "duration" in c_low or "dur" in c_low:
                    mapping["duration"] = col
                elif "time" in c_low or "actual" in c_low:
                    mapping["time"] = col

        # Fallbacks
        if "id" not in mapping and len(df.columns) > 0:
            mapping["id"] = df.columns[0]
        if "break" not in mapping and len(df.columns) > 1:
            mapping["break"] = df.columns[1]
        if "duration" not in mapping and len(df.columns) > 2:
            mapping["duration"] = df.columns[2]
        if "time" not in mapping:
            mapping["time"] = df.columns[-1]

        return mapping

    def populate_table(self, table, data, headers, bg_color):
        table.setRowCount(0)
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)

        for row_idx, row_data in enumerate(data):
            table.insertRow(row_idx)
            for col_idx, value in enumerate(row_data):
                item = QTableWidgetItem(str(value))
                item.setBackground(bg_color)
                item.setForeground(QColor("#FFFFFF"))
                item.setFont(QFont("Arial", 10))
                table.setItem(row_idx, col_idx, item)

    def filter_tables(self):
        query = self.txt_search.text().lower()
        for table in [self.tbl_played, self.tbl_missed, self.tbl_extra, self.tbl_mismatch]:
            for row in range(table.rowCount()):
                match = False
                for col in range(table.columnCount()):
                    item = table.item(row, col)
                    if item and query in item.text().lower():
                        match = True
                        break
                table.setRowHidden(row, not match)

    def export_data(self, file_format):
        if not self.results["played"] and not self.results["missed"]:
            QMessageBox.warning(self, "No Data", "No verified results to export.")
            return

        path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Report",
            f"Spot_Verification_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{file_format}",
            "Excel Files (*.xlsx)" if file_format == "xlsx" else "CSV Files (*.csv)"
        )

        if not path:
            return

        try:
            # Flatten results
            all_records = []
            status_color = {
                "PLAYED": "#1B5E20",
                "MISSED": "#B71C1C",
                "EXTRA": "#0D47A1",
                "MISMATCH": "#F57F17"
            }

            for status, rows in self.results.items():
                status_label = status.upper().replace("_", " ")
                for r in rows:
                    all_records.append([status_label] + r)

            columns = ["Status", "Creative ID", "Break", "Slot", "Sched Dur", "Actual Dur", "Brand", "Air Time"]
            final_df = pd.DataFrame(all_records, columns=columns)

            if file_format == "xlsx":
                # Create colored Excel report
                wb = Workbook()
                ws = wb.active
                ws.title = "Verification Report"

                # Headers
                header_fill = PatternFill(start_color="0F2640", end_color="0F2640", fill_type="solid")
                header_font = Font(color="00D4FF", bold=True, size=12)

                for col_idx, col_name in enumerate(columns, 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.value = col_name
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Data rows with color coding
                color_map = {
                    "PLAYED": "C6EFCE",
                    "MISSED": "FFC7CE",
                    "EXTRA": "BDD7EE",
                    "MISMATCH": "FFEB9C"
                }

                for row_idx, row_data in enumerate(final_df.values, 2):
                    status = row_data[0]
                    color = color_map.get(status, "FFFFFF")
                    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = value
                        cell.fill = fill
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = Border(
                            left=Side(style="thin"),
                            right=Side(style="thin"),
                            top=Side(style="thin"),
                            bottom=Side(style="thin")
                        )

                # Adjust column widths
                for col in ws.columns:
                    max_length = 0
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[cell.column_letter].width = min(max_length + 2, 30)

                wb.save(path)
            else:
                final_df.to_csv(path, index=False)

            QMessageBox.information(self, "✅ Export Successful", f"Report saved to:\n{path}")

        except Exception as e:
            QMessageBox.critical(self, "❌ Export Failed", f"Error: {str(e)}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = BroadcastSpotVerifier()
    window.show()
    sys.exit(app.exec())
